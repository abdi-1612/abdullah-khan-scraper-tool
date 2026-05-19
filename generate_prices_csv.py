#!/usr/bin/env python3
"""
Generate a workbook-driven prices.csv for the local deal finder.

The workbook stores source prices in USD. This script keeps `base_price_usd`
as the source of truth, while also exporting derived alert thresholds so the
CSV stays human-readable:
- sell price = 1.4 x sheet USD price
- max buy = 1.12 x sheet USD price
- alert ceiling = max buy x negotiation buffer
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional

SELL_PRICE_MULTIPLIER = 1.4
MAX_BUY_MULTIPLIER = 1.12
NEGOTIATION_ABOVE_MAX_BUY_RATIO = 1.10
MAX_ALERT_MULTIPLIER = MAX_BUY_MULTIPLIER * NEGOTIATION_ABOVE_MAX_BUY_RATIO

DEFAULT_WORKBOOK = Path("/Users/Abdullah/Downloads/Atlas Mobile Price Sheet.xlsx")
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "prices.csv"

FIELDNAMES = [
    "category",
    "brand",
    "family_key",
    "model",
    "model_year",
    "device_line",
    "chip_family",
    "storage_options",
    "carrier_status",
    "default_condition",
    "source_sheet",
    "base_price_usd",
    "fair_price_from_sheet_cad",
    "maximum_buy_price_cad",
    "max_listing_price_to_alert_cad",
    "condition_prices_usd",
    "deduction_rules_usd",
    "aliases",
    "notes",
]


def json_compact(value: object) -> str:
    return json.dumps(value, sort_keys=True, separators=(",", ":"))


def normalize_space(text: str) -> str:
    return " ".join(str(text).replace("\xa0", " ").split())


def normalize_text(text: str) -> str:
    text = normalize_space(text).lower()
    text = text.replace("&", " and ")
    text = text.replace('"', " inch ")
    text = text.replace("'", " ")
    text = text.replace("+", " plus ")
    text = text.replace("/", " / ")
    text = text.replace("se 2020", "se 2nd gen")
    text = text.replace("se 2022", "se 3rd gen")
    text = text.replace("second generation", "2nd gen")
    text = text.replace("third generation", "3rd gen")
    text = text.replace("second gen", "2nd gen")
    text = text.replace("third gen", "3rd gen")
    text = re.sub(r"([a-z])(\d)", r"\1 \2", text)
    text = re.sub(r"(\d)([a-z])", r"\1 \2", text)
    text = text.replace("gb", " gb")
    text = text.replace("tb", " tb")
    text = re.sub(r"[^a-z0-9+/ ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def slug(text: str) -> str:
    return normalize_text(text).replace(" / ", "_").replace(" ", "_")


def as_number(value: object) -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if stripped.upper() in {"ASK", "NOT BUYING", "NOT BUYING ", "NOT BUYING", "#REF!", "#NUM!", "-", "USED"}:
            return None
        stripped = stripped.replace("$", "").replace(",", "")
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def sell_price_from_usd(value: float) -> float:
    return round(value * SELL_PRICE_MULTIPLIER, 2)


def max_buy_price_from_usd(value: float) -> float:
    return round(value * MAX_BUY_MULTIPLIER, 2)


def alert_ceiling_from_usd(value: float) -> float:
    return round(value * MAX_ALERT_MULTIPLIER, 2)


def storage_sort_key(storage: str) -> int:
    normalized = normalize_text(storage)
    match = re.match(r"(\d+)\s*(gb|tb|mm)", normalized)
    if not match:
        return 999999
    amount = int(match.group(1))
    unit = match.group(2)
    if unit == "tb":
        return amount * 1000
    if unit == "mm":
        return amount
    return amount


def normalize_storage_option(option: str) -> str:
    normalized = normalize_text(option)
    normalized = normalized.replace(" gb", "gb").replace(" tb", "tb").replace(" mm", "mm")
    return normalized


def split_storage_options(text: str) -> List[str]:
    if not text:
        return []
    options = []
    for raw in re.split(r"/", text):
        raw = raw.strip()
        if not raw:
            continue
        options.append(normalize_storage_option(raw))
    return options


def normalize_samsung_model_name(raw: str) -> str:
    normalized = normalize_text(raw)
    if not normalized:
        return ""

    if any(
        phrase in normalized
        for phrase in (
            "must be",
            "google locked",
            "kg status",
            "trade in lock",
            "lcd burns",
            "shadow",
            "check for",
            "responsible",
            "missing stylus",
            "deductions please ask",
        )
    ):
        return ""

    compact = normalized.replace("samsung ", "").replace("galaxy ", "").strip()

    series_match = re.fullmatch(r"s\s*(\d{2})\s*(ultra|plus|edge|fe)?", compact)
    if series_match:
        number = series_match.group(1)
        variant = series_match.group(2)
        parts = [f"Galaxy S{number}"]
        if variant == "plus":
            parts.append("Plus")
        elif variant == "ultra":
            parts.append("Ultra")
        elif variant == "edge":
            parts.append("Edge")
        elif variant == "fe":
            parts.append("FE")
        return " ".join(parts)

    note_match = re.fullmatch(r"note\s*(\d{2})\s*(ultra|plus)?", compact)
    if note_match:
        number = note_match.group(1)
        variant = note_match.group(2)
        parts = [f"Galaxy Note {number}"]
        if variant == "ultra":
            parts.append("Ultra")
        elif variant == "plus":
            parts.append("Plus")
        return " ".join(parts)

    z_flip_match = re.fullmatch(r"(?:z\s+)?flip\s*(\d)", compact)
    if z_flip_match:
        return f"Galaxy Z Flip {z_flip_match.group(1)}"

    z_fold_match = re.fullmatch(r"(?:z\s+)?fold\s*(\d)", compact)
    if z_fold_match:
        return f"Galaxy Z Fold {z_fold_match.group(1)}"

    fold_match = re.fullmatch(r"fold\s*(\d)", compact)
    if fold_match:
        return f"Galaxy Fold {fold_match.group(1)}"

    if normalized.startswith("galaxy "):
        return normalize_space(raw)

    return ""


def parse_deduction_notes(text: str) -> Dict[str, dict]:
    raw_lower = normalize_space(text).lower()
    raw_lower = raw_lower.replace("—", "-").replace("–", "-").replace("−", "-")
    raw_lower = re.sub(r"\s+", " ", raw_lower).strip()
    lower = normalize_text(text)
    rules: Dict[str, dict] = {}

    segments: List[str] = []
    for chunk in re.split(r"/|\|", raw_lower):
        parts = re.split(r"\s*-\s*-\s*-\s*", chunk)
        for part in parts:
            normalized = normalize_space(part).lower()
            if normalized:
                segments.append(normalized)

    def extract_amount(segment: str) -> Optional[float]:
        match = re.search(r"(?:=|is)?\s*-?\s*\$?\s*(\d+(?:\.\d+)?)\s*(?:off)?\b", segment)
        if not match:
            return None
        return float(match.group(1))

    def add_flat_rule(key: str, amount: Optional[float]) -> None:
        if amount is not None:
            rules[key] = {"kind": "flat_usd", "value": amount}

    def add_rule_from_segment(key: str, segment: str) -> None:
        amount = extract_amount(segment)
        if amount is not None:
            rules[key] = {"kind": "flat_usd", "value": amount}
            return
        if "parts" in segment:
            rules[key] = {"kind": "manual", "reason": "parts"}
            return
        grade_match = re.search(r"\bgrade\s*(c|d|doa)\b", segment)
        if grade_match:
            rules[key] = {"kind": "condition_override", "value": grade_match.group(1)}
            return
        if "ask" in segment:
            rules[key] = {"kind": "manual", "reason": "ask"}

    for segment in segments:
        if segment.startswith("cracked back"):
            add_flat_rule("cracked_back", extract_amount(segment))
        elif segment.startswith("cracked lens"):
            add_flat_rule("cracked_lens", extract_amount(segment))
        elif segment.startswith("google locked"):
            add_flat_rule("google_locked", extract_amount(segment))
        elif segment.startswith("no charger"):
            add_flat_rule("no_charger", extract_amount(segment))
        elif segment.startswith("no band"):
            add_flat_rule("no_band", extract_amount(segment))
        elif segment.startswith("ultra missing band"):
            add_flat_rule("ultra_missing_band", extract_amount(segment))
        elif segment.startswith("open box"):
            add_flat_rule("open_box", extract_amount(segment))
        elif segment.startswith("bad face id"):
            add_rule_from_segment("bad_face_id", segment)
        elif segment.startswith("bad back camera"):
            add_rule_from_segment("bad_back_camera", segment)

    if "degraded battery" in lower or "battery under 80" in lower:
        rules["battery_under_80"] = {"kind": "manual", "reason": "extra_deduction"}
    if "repair message" in lower:
        rules["repair_message"] = {"kind": "manual", "reason": "extra_deduction"}
    if "heavy scratching" in lower:
        rules["heavy_scratching"] = {"kind": "manual", "reason": "extra_deduction"}
    if "lcd burns" in lower or "lcd burn" in lower or "shadow" in lower:
        rules["lcd_burn"] = {"kind": "manual", "reason": "ask"}
    if "kg status active" in lower or "trade in lock" in lower or "priced as parts" in lower:
        rules["kg_active"] = {"kind": "manual", "reason": "parts"}
    if "engraving" in lower:
        rules["engraved"] = {"kind": "manual", "reason": "ask"}
    if "missing parts" in lower or "must be complete" in lower:
        rules["missing_parts"] = {"kind": "manual", "reason": "ask"}
    if "demo" in lower and "ask" in lower:
        rules["demo"] = {"kind": "manual", "reason": "ask"}

    return rules


def choose_default_condition(
    condition_prices: Dict[str, float],
    preferred_order: Iterable[str],
) -> tuple[str, Optional[float]]:
    for key in preferred_order:
        if key in condition_prices:
            return key, condition_prices[key]
    if condition_prices:
        key = sorted(condition_prices.keys())[0]
        return key, condition_prices[key]
    return "", None


def join_storage_options(options: List[str]) -> str:
    if not options:
        return ""
    unique = []
    seen = set()
    for option in sorted(options, key=storage_sort_key):
        if option not in seen:
            seen.add(option)
            unique.append(option)
    return "|".join(unique)


def lowest_storage_value(options: List[str]) -> str:
    if not options:
        return ""
    return sorted(options, key=storage_sort_key)[0]


def is_useful_alias(alias: str) -> bool:
    tokens = [token for token in normalize_text(alias).split() if token]
    if not tokens:
        return False

    if len(tokens) == 1 and tokens[0].isdigit():
        return False

    joined = "".join(tokens)
    if len(joined) < 4:
        return False

    return True


def build_aliases(category: str, model: str, storage_options: List[str], carrier_status: str) -> List[str]:
    normalized_model = normalize_text(model)
    aliases = {normalized_model}

    brandless = normalized_model
    for prefix in ["apple ", "iphone ", "ipad ", "macbook ", "galaxy ", "google ", "pixel ", "watch "]:
        if brandless.startswith(prefix):
            stripped = brandless[len(prefix) :].strip()
            if len(stripped.split()) >= 2:
                aliases.add(stripped)

    if category == "iphone":
        aliases.add(normalized_model.replace("iphone ", "").strip())
        aliases.add(normalized_model.replace("pro max", "pm"))
        aliases.add(normalized_model.replace(" plus", " plus"))
        if "se 2nd gen" in normalized_model:
            aliases.add("iphone se 2020")
            aliases.add("se 2020")
        if "se 3rd gen" in normalized_model:
            aliases.add("iphone se 2022")
            aliases.add("se 2022")

    if category == "samsung":
        aliases.add(normalized_model.replace("galaxy ", "").strip())

    if category == "pixel":
        aliases.add(normalized_model.replace("google ", "").replace("pixel ", "").strip())
        aliases.add(normalized_model.replace("google ", "").strip())

    if category == "ipad":
        aliases.add(normalized_model.replace("apple ", "").strip())
        aliases.add(normalized_model.replace("ipad ", "").strip())

    if category == "macbook":
        aliases.add(normalized_model.replace("macbook pro", "mbp"))
        aliases.add(normalized_model.replace("macbook air", "mba"))

    if category == "apple_watch":
        aliases.add(normalized_model.replace("apple watch ", "").strip())
        aliases.add(normalized_model.replace("apple watch ", "watch ").strip())

    for storage in storage_options:
        aliases.add(f"{normalized_model} {storage}")
        aliases.add(f"{brandless} {storage}")

    if carrier_status == "carrier_locked":
        aliases.add(f"{normalized_model} locked")
        aliases.add(f"{normalized_model} carrier locked")
    elif carrier_status == "unlocked":
        aliases.add(f"{normalized_model} unlocked")

    cleaned = []
    seen = set()
    for alias in aliases:
        alias = normalize_text(alias)
        if not alias or alias in seen or not is_useful_alias(alias):
            continue
        seen.add(alias)
        cleaned.append(alias)
    return sorted(cleaned)


def row_record(
    *,
    category: str,
    brand: str,
    model: str,
    model_year: str = "",
    device_line: str = "",
    chip_family: str = "",
    storage_options: List[str],
    carrier_status: str,
    default_condition: str,
    base_price_usd: float,
    condition_prices_usd: Dict[str, float],
    deduction_rules_usd: Dict[str, dict],
    source_sheet: str,
    notes: str = "",
) -> dict:
    storage_string = join_storage_options(storage_options)
    family_key = f"{slug(model)}__{carrier_status or 'na'}"
    aliases = build_aliases(category, model, storage_options, carrier_status)
    return {
        "category": category,
        "brand": brand,
        "family_key": family_key,
        "model": normalize_text(model),
        "model_year": normalize_text(model_year),
        "device_line": normalize_text(device_line),
        "chip_family": normalize_text(chip_family),
        "storage_options": storage_string,
        "carrier_status": carrier_status,
        "default_condition": default_condition,
        "source_sheet": source_sheet,
        "base_price_usd": f"{base_price_usd:.2f}",
        "fair_price_from_sheet_cad": f"{sell_price_from_usd(base_price_usd):.2f}",
        "maximum_buy_price_cad": f"{max_buy_price_from_usd(base_price_usd):.2f}",
        "max_listing_price_to_alert_cad": f"{alert_ceiling_from_usd(base_price_usd):.2f}",
        "condition_prices_usd": json_compact(condition_prices_usd),
        "deduction_rules_usd": json_compact(deduction_rules_usd),
        "aliases": "|".join(aliases),
        "notes": normalize_space(notes),
    }


def merge_rows(rows: List[dict]) -> List[dict]:
    merged: Dict[tuple, dict] = {}
    for row in rows:
        key = (
            row["category"],
            row["model"],
            row["storage_options"],
            row["carrier_status"],
        )
        if key not in merged:
            merged[key] = row
            continue

        current = merged[key]
        current["base_price_usd"] = f"{min(float(current['base_price_usd']), float(row['base_price_usd'])):.2f}"
        current["fair_price_from_sheet_cad"] = f"{min(float(current['fair_price_from_sheet_cad']), float(row['fair_price_from_sheet_cad'])):.2f}"
        current["maximum_buy_price_cad"] = f"{min(float(current['maximum_buy_price_cad']), float(row['maximum_buy_price_cad'])):.2f}"
        current["max_listing_price_to_alert_cad"] = f"{min(float(current['max_listing_price_to_alert_cad']), float(row['max_listing_price_to_alert_cad'])):.2f}"

        left_prices = json.loads(current["condition_prices_usd"])
        right_prices = json.loads(row["condition_prices_usd"])
        for label, value in right_prices.items():
            if label not in left_prices:
                left_prices[label] = value
            else:
                left_prices[label] = min(left_prices[label], value)
        current["condition_prices_usd"] = json_compact(left_prices)

        left_rules = json.loads(current["deduction_rules_usd"])
        right_rules = json.loads(row["deduction_rules_usd"])
        left_rules.update(right_rules)
        current["deduction_rules_usd"] = json_compact(left_rules)

        alias_set = set(filter(None, current["aliases"].split("|")))
        alias_set.update(filter(None, row["aliases"].split("|")))
        current["aliases"] = "|".join(sorted(alias_set))

        if row["notes"] and row["notes"] not in current["notes"]:
            current["notes"] = " | ".join(part for part in [current["notes"], row["notes"]] if part)

    def sort_key(row: dict) -> tuple:
        storage_key = storage_sort_key(lowest_storage_value(row["storage_options"].split("|")))
        return (row["category"], row["model"], row["carrier_status"], storage_key)

    return sorted(merged.values(), key=sort_key)


def parse_iphone_rows(workbook) -> List[dict]:
    ws = workbook["iPhone Used"]
    rows = []
    current_notes = ""
    current_deductions: Dict[str, dict] = {}
    model_pattern = re.compile(
        r"^(?P<model>(?:iPhone\s+.+?|SE \(\d{4}\)))\s+(?P<storage>\d+(?:GB|TB))\s+(?P<status>Unlocked|Carrier Locked)$",
        re.IGNORECASE,
    )

    for row_idx in range(1, ws.max_row + 1):
        raw = ws[f"B{row_idx}"].value
        if not isinstance(raw, str):
            continue
        text = normalize_space(raw)
        lower = normalize_text(text)

        if "cracked back" in lower or "degraded battery" in lower or "repair message" in lower:
            current_notes = text
            current_deductions = parse_deduction_notes(text)
            continue

        match = model_pattern.match(text)
        if not match:
            continue

        model = match.group("model")
        if model.lower().startswith("se (2020)"):
            model = "iPhone SE (2nd Gen)"

        condition_prices = {
            "a": as_number(ws[f"D{row_idx}"].value),
            "b": as_number(ws[f"E{row_idx}"].value),
            "c": as_number(ws[f"F{row_idx}"].value),
            "d": as_number(ws[f"G{row_idx}"].value),
            "doa": as_number(ws[f"H{row_idx}"].value),
        }
        condition_prices = {k: v for k, v in condition_prices.items() if v is not None and v > 0}
        default_condition, base_price = choose_default_condition(condition_prices, ["b", "a", "c", "d"])
        if base_price is None:
            continue

        rows.append(
            row_record(
                category="iphone",
                brand="apple",
                model=model,
                storage_options=[normalize_storage_option(match.group("storage"))],
                carrier_status="unlocked" if match.group("status").lower() == "unlocked" else "carrier_locked",
                default_condition=default_condition,
                base_price_usd=base_price,
                condition_prices_usd=condition_prices,
                deduction_rules_usd=current_deductions,
                source_sheet="iPhone Used",
                notes=current_notes,
            )
        )
    return rows


def split_model_storage_suffix(raw: str) -> tuple[str, List[str]]:
    match = re.search(r"(.+?)\s+((?:\d+(?:GB|TB))(?:/\d+(?:GB|TB))*)$", normalize_space(raw), re.IGNORECASE)
    if not match:
        return normalize_space(raw), []
    return match.group(1), split_storage_options(match.group(2))


def parse_ipad_rows(workbook) -> List[dict]:
    ws = workbook["iPad Master"]
    rows = []
    deductions = parse_deduction_notes("ENGRAVINGS = ASK")
    for row_idx in range(1, ws.max_row + 1):
        model_cell = ws[f"B{row_idx}"].value
        grade_b = as_number(ws[f"D{row_idx}"].value)
        if not isinstance(model_cell, str) or grade_b is None or grade_b <= 0:
            continue
        model_raw, storage_options = split_model_storage_suffix(model_cell)
        condition_prices = {
            "a": as_number(ws[f"C{row_idx}"].value),
            "b": grade_b,
            "c": as_number(ws[f"E{row_idx}"].value),
            "d": as_number(ws[f"F{row_idx}"].value),
            "doa": as_number(ws[f"G{row_idx}"].value),
        }
        condition_prices = {k: v for k, v in condition_prices.items() if v is not None and v > 0}
        default_condition, base_price = choose_default_condition(condition_prices, ["b", "a", "c", "d"])
        if base_price is None:
            continue
        rows.append(
            row_record(
                category="ipad",
                brand="apple",
                model=model_raw,
                storage_options=storage_options,
                carrier_status="not_applicable",
                default_condition=default_condition,
                base_price_usd=base_price,
                condition_prices_usd=condition_prices,
                deduction_rules_usd=deductions,
                source_sheet="iPad Master",
                notes="Used pricing. Engravings require manual review.",
            )
        )
    return rows


def parse_samsung_rows(workbook) -> List[dict]:
    ws = workbook["Samsung Master"]
    rows = []
    deductions = {}
    for row_idx in range(1, 15):
        cell = ws[f"B{row_idx}"].value
        if isinstance(cell, str):
            deductions.update(parse_deduction_notes(cell))

    current_model = ""
    current_note = ""
    for row_idx in range(1, ws.max_row + 1):
        model_cell = ws[f"B{row_idx}"].value
        status_cell = ws[f"C{row_idx}"].value
        model_text = normalize_space(model_cell) if isinstance(model_cell, str) else ""
        normalized_model = normalize_samsung_model_name(model_text)
        numeric_values = [as_number(ws[f"{col}{row_idx}"].value) for col in "DEFGHI"]
        has_prices = any(value is not None and value > 0 for value in numeric_values)
        if not has_prices:
            if normalized_model:
                current_model = normalized_model
                current_note = ""
            elif model_text and any(
                phrase in normalize_text(model_text)
                for phrase in ("missing stylus", "deductions please ask")
            ):
                current_note = model_text
            continue

        model = normalized_model or current_model
        if normalized_model:
            current_model = normalized_model
        if not model:
            continue

        carrier_status = "not_applicable"
        if isinstance(status_cell, str):
            normalized_status = normalize_text(status_cell)
            if "carrier locked" in normalized_status:
                carrier_status = "carrier_locked"
            elif "unlocked" in normalized_status:
                carrier_status = "unlocked"

        condition_prices = {
            "new": as_number(ws[f"D{row_idx}"].value),
            "a": as_number(ws[f"E{row_idx}"].value),
            "b": as_number(ws[f"F{row_idx}"].value),
            "c": as_number(ws[f"G{row_idx}"].value),
            "d": as_number(ws[f"H{row_idx}"].value),
            "doa": as_number(ws[f"I{row_idx}"].value),
        }
        condition_prices = {k: v for k, v in condition_prices.items() if v is not None and v > 0}
        default_condition, base_price = choose_default_condition(condition_prices, ["b", "a", "c", "d", "new"])
        if base_price is None:
            continue

        rows.append(
            row_record(
                category="samsung",
                brand="samsung",
                model=model,
                storage_options=[],
                carrier_status=carrier_status,
                default_condition=default_condition,
                base_price_usd=base_price,
                condition_prices_usd=condition_prices,
                deduction_rules_usd=deductions,
                source_sheet="Samsung Master",
                notes=" | ".join(
                    part
                    for part in [
                        "Samsung sheet uses condition columns with some ASK / not buying gaps.",
                        current_note,
                    ]
                    if part
                ),
            )
        )
    return rows


def parse_pixel_model_name(raw: str) -> str:
    raw = normalize_text(raw)
    raw_compact = raw.replace(" ", "")
    replacements = {
        "pixel pro fold 2024 5g factory original unlocked": "google pixel 9 pro fold",
        "pixel 10 pro fold 2025 5g factory original unlocked": "google pixel 10 pro fold",
        "pixel 10 pro xl 6 8 5g factory original unlocked": "google pixel 10 pro xl",
        "pixel 10 pro 6 3 5g factory original unlocked": "google pixel 10 pro",
        "pixel 10 5g factory original unlocked": "google pixel 10",
        "pixel 9 pro xl 6 8 5g factory original unlocked": "google pixel 9 pro xl",
        "pixel 9 pro 6 5 5g factory original unlocked": "google pixel 9 pro",
        "pixel 9 5g factory original unlocked 9a 50": "google pixel 9",
        "pixel 8 5g factory original unlocked pixel 8a 50": "google pixel 8",
        "google pixel fold new 2023 5g": "google pixel fold",
        "pixel 8 pro 5g factory original unlocked": "google pixel 8 pro",
        "pixel 7 pro 5g factory original unlocked": "google pixel 7 pro",
        "pixel 7 5g factory original unlocked": "google pixel 7",
        "pixel 6 pro 5g factory original unlocked": "google pixel 6 pro",
        "pixel 6 5g factory original unlocked": "google pixel 6",
        "pixel 6a 5g factory original unlocked": "google pixel 6a",
        "pixel 5a 5g factory original unlocked": "google pixel 5a",
        "pixel 5 5g factory original unlocked": "google pixel 5",
        "10pro fold": "google pixel 10 pro fold",
        "10 pro xl": "google pixel 10 pro xl",
        "10 pro": "google pixel 10 pro",
        "10": "google pixel 10",
        "9 fold": "google pixel 9 fold",
        "9 pro xl": "google pixel 9 pro xl",
        "9 pro": "google pixel 9 pro",
        "9": "google pixel 9",
        "9a": "google pixel 9a",
        "9 a": "google pixel 9a",
        "8 fold": "google pixel 8 fold",
        "8 pro": "google pixel 8 pro",
        "8": "google pixel 8",
        "8a": "google pixel 8a",
        "8 a": "google pixel 8a",
        "7 pro": "google pixel 7 pro",
        "7": "google pixel 7",
        "7a": "google pixel 7a",
        "7 a": "google pixel 7a",
        "6pro": "google pixel 6 pro",
        "6": "google pixel 6",
        "6a": "google pixel 6a",
        "6 a": "google pixel 6a",
    }
    return replacements.get(raw, replacements.get(raw_compact, raw))


def add_pixel_adjusted_variant(
    rows: List[dict],
    *,
    model: str,
    carrier_status: str,
    default_condition: str,
    condition_prices: Dict[str, float],
    deduction_rules: Dict[str, dict],
    source_sheet: str,
    notes: str,
    adjust_name: str,
    adjust_amount: float,
) -> None:
    adjusted_prices = {}
    for label, value in condition_prices.items():
        adjusted_value = round(value - adjust_amount, 2)
        if adjusted_value > 0:
            adjusted_prices[label] = adjusted_value
    default_condition, base_price = choose_default_condition(
        adjusted_prices,
        [default_condition, "b_plus", "ab_grade", "a", "c", "d", "open", "sealed", "new"],
    )
    if base_price is None:
        return
    rows.append(
        row_record(
            category="pixel",
            brand="google",
            model=adjust_name,
            storage_options=[],
            carrier_status=carrier_status,
            default_condition=default_condition,
            base_price_usd=base_price,
            condition_prices_usd=adjusted_prices,
            deduction_rules_usd=deduction_rules,
            source_sheet=source_sheet,
            notes=notes,
        )
    )


def parse_pixel_rows(workbook) -> List[dict]:
    ws = workbook["Pixel Master"]
    rows = []

    for row_idx in range(3, ws.max_row + 1):
        raw = ws[f"A{row_idx}"].value
        if not isinstance(raw, str):
            continue
        model = parse_pixel_model_name(raw)
        if not model.startswith("google pixel"):
            continue

        condition_prices = {
            "sealed": as_number(ws[f"B{row_idx}"].value),
            "open": as_number(ws[f"C{row_idx}"].value),
            "a": as_number(ws[f"D{row_idx}"].value),
            "b_plus": as_number(ws[f"E{row_idx}"].value),
        }
        condition_prices = {k: v for k, v in condition_prices.items() if v is not None and v > 0}
        default_condition, base_price = choose_default_condition(condition_prices, ["b_plus", "a", "open", "sealed"])
        if base_price is None:
            continue
        deductions = {"damage_or_demo": {"kind": "manual", "reason": "ask"}}
        rows.append(
            row_record(
                category="pixel",
                brand="google",
                model=model,
                storage_options=[],
                carrier_status="unlocked",
                default_condition=default_condition,
                base_price_usd=base_price,
                condition_prices_usd=condition_prices,
                deduction_rules_usd=deductions,
                source_sheet="Pixel Master",
                notes="Unlocked Pixel sheet uses B+ as the default used condition.",
            )
        )

        if model == "google pixel 9":
            add_pixel_adjusted_variant(
                rows,
                model=model,
                carrier_status="unlocked",
                default_condition=default_condition,
                condition_prices=condition_prices,
                deduction_rules=deductions,
                source_sheet="Pixel Master",
                notes="Atlas notes 9A is $50 below Pixel 9 unlocked row.",
                adjust_name="google pixel 9a",
                adjust_amount=50,
            )
        if model == "google pixel 8":
            add_pixel_adjusted_variant(
                rows,
                model=model,
                carrier_status="unlocked",
                default_condition=default_condition,
                condition_prices=condition_prices,
                deduction_rules=deductions,
                source_sheet="Pixel Master",
                notes="Atlas notes 8A is $50 below Pixel 8 unlocked row.",
                adjust_name="google pixel 8a",
                adjust_amount=50,
            )

    for row_idx in range(24, ws.max_row + 1):
        raw = ws[f"A{row_idx}"].value
        if raw in (None, ""):
            continue
        model = parse_pixel_model_name(str(raw))
        if not model.startswith("google pixel"):
            continue

        condition_prices = {
            "new": as_number(ws[f"B{row_idx}"].value),
            "ab_grade": as_number(ws[f"C{row_idx}"].value),
            "c": as_number(ws[f"D{row_idx}"].value),
            "d": as_number(ws[f"E{row_idx}"].value),
            "doa": as_number(ws[f"F{row_idx}"].value),
        }
        condition_prices = {k: v for k, v in condition_prices.items() if v is not None and v > 0}
        default_condition, base_price = choose_default_condition(condition_prices, ["ab_grade", "c", "d", "new"])
        if base_price is None:
            continue

        deductions = {}
        cracked_back = as_number(ws[f"G{row_idx}"].value)
        demo = as_number(ws[f"H{row_idx}"].value)
        if cracked_back is not None and cracked_back < 0:
            deductions["cracked_back"] = {"kind": "flat_usd", "value": abs(cracked_back)}
        if demo is not None and demo < 0:
            deductions["demo"] = {"kind": "flat_usd", "value": abs(demo)}

        rows.append(
            row_record(
                category="pixel",
                brand="google",
                model=model,
                storage_options=[],
                carrier_status="carrier_locked",
                default_condition=default_condition,
                base_price_usd=base_price,
                condition_prices_usd=condition_prices,
                deduction_rules_usd=deductions,
                source_sheet="Pixel Master",
                notes="Carrier-locked Pixel sheet uses A/B grade as the default used condition.",
            )
        )
    return rows


def parse_watch_rows(workbook) -> List[dict]:
    ws = workbook["Apple Watch"]
    rows = []
    deductions = {}
    for row_idx in range(1, 6):
        cell = ws[f"B{row_idx}"].value
        if isinstance(cell, str):
            deductions.update(parse_deduction_notes(cell))
    for row_idx in range(1, ws.max_row + 1):
        model_cell = ws[f"B{row_idx}"].value
        b_grade = as_number(ws[f"F{row_idx}"].value)
        if not isinstance(model_cell, str) or b_grade is None or b_grade <= 0:
            continue
        model = f"Apple Watch {normalize_space(model_cell)}"
        condition_prices = {
            "sealed": as_number(ws[f"C{row_idx}"].value),
            "open": as_number(ws[f"D{row_idx}"].value),
            "a": as_number(ws[f"E{row_idx}"].value),
            "b": b_grade,
        }
        condition_prices = {k: v for k, v in condition_prices.items() if v is not None and v > 0}
        default_condition, base_price = choose_default_condition(condition_prices, ["b", "a", "open", "sealed"])
        if base_price is None:
            continue

        note_cell = ws[f"G{row_idx}"].value
        note_text = normalize_space(note_cell) if isinstance(note_cell, str) else ""
        rows.append(
            row_record(
                category="apple_watch",
                brand="apple",
                model=model,
                storage_options=[],
                carrier_status="not_applicable",
                default_condition=default_condition,
                base_price_usd=base_price,
                condition_prices_usd=condition_prices,
                deduction_rules_usd=deductions,
                source_sheet="Apple Watch",
                notes=note_text,
            )
        )
    return rows


MACBOOK_OVERRIDES = {
    ("2025 Macbook Pro 14 inch M5", "M5 Pro MDE04 512GB"): ("MacBook Pro 14 M5 Pro", ["512GB"], "Base 2025 M5 Pro 14-inch"),
    ("2025 Macbook Pro 14 inch M5", "M5 Pro MDE14 1TB"): ("MacBook Pro 14 M5 Pro", ["1TB"], "Higher-storage 2025 M5 Pro 14-inch"),
    ("2025 Macbook Pro 14 inch M5", "M5 MAX MDE34 1TB"): ("MacBook Pro 14 M5 Max", ["1TB"], "2025 M5 Max 14-inch"),
    ("2025 MacBook Air 15 inch M4", "MC7A4 MW1G3 MW1L3 MW1J3"): ("MacBook Air 15 M4", ["256GB"], "Part-code row inferred as the lowest-storage M4 Air 15-inch variant"),
    ("2025 MacBook Air 15 inch M4", "MC7C4 MW1H3 MW1M3 MW1K3"): ("MacBook Air 15 M4", ["512GB"], "Part-code row inferred as the mid-storage M4 Air 15-inch variant"),
    ("2025 MacBook Air 15 inch M4", "MC7D4 MC6J4 MC6L4 MC6K4"): ("MacBook Air 15 M4 24GB", ["512GB"], "Part-code row inferred as the higher-memory M4 Air 15-inch variant"),
    ("2025 MacBook Air 13 inch M4", "MC6T4 MW0W3 MW123 MW0Y3"): ("MacBook Air 13 M4", ["256GB"], "Part-code row inferred as the lowest-storage M4 Air 13-inch variant"),
    ("2025 MacBook Air 13 inch M4", "MC6U4 MW0X3 MW133 MW103"): ("MacBook Air 13 M4", ["512GB"], "Part-code row inferred as the mid-storage M4 Air 13-inch variant"),
    ("2025 MacBook Air 13 inch M4", "MC6V4 MC654 MC6C4 MC6A4"): ("MacBook Air 13 M4 24GB", ["512GB"], "Part-code row inferred as the higher-memory M4 Air 13-inch variant"),
    ("2024 Macbook Pro 16 inch M4", "M4 Pro MX2X3 MX2T3 512GB"): ("MacBook Pro 16 M4 Pro", ["512GB"], "One of multiple 16-inch M4 Pro 512GB rows"),
    ("2024 Macbook Pro 16 inch M4", "M4 Pro MX2Y3 MX2U3 512GB"): ("MacBook Pro 16 M4 Pro High", ["512GB"], "Higher-spec 16-inch M4 Pro 512GB row"),
    ("2024 Macbook Pro 16 inch M4", "M4 MAX MX303 MX2V3 1TB"): ("MacBook Pro 16 M4 Max", ["1TB"], "Lower of the visible 16-inch M4 Max 1TB rows"),
    ("2024 Macbook Pro 16 inch M4", "M4 MAX MX313 MX2W3 1TB"): ("MacBook Pro 16 M4 Max High", ["1TB"], "Higher-spec 16-inch M4 Max 1TB row"),
    ("2024 Macbook Pro 14 inch M4", "MW2U3 MW2W3 512GB"): ("MacBook Pro 14 M4", ["512GB"], "Base 14-inch M4 row"),
    ("2024 Macbook Pro 14 inch M4", "MW2V3 MW2X3 1TB"): ("MacBook Pro 14 M4", ["1TB"], "Base 14-inch M4 1TB row"),
    ("2024 Macbook Pro 14 inch M4", "MCX04 MCX14 1TB"): ("MacBook Pro 14 M4 High", ["1TB"], "Higher-spec 14-inch M4 1TB row"),
    ("2024 Macbook Pro 14 inch M4", "M4 Pro MX2H3 MX2E3 512GB"): ("MacBook Pro 14 M4 Pro", ["512GB"], "14-inch M4 Pro 512GB row"),
    ("2024 Macbook Pro 14 inch M4", "M4 Pro MX2J3 MX2F3 1TB"): ("MacBook Pro 14 M4 Pro", ["1TB"], "14-inch M4 Pro 1TB row"),
    ("2024 Macbook Pro 14 inch M4", "M4 MAX MX2K3 MX2G3 1TB"): ("MacBook Pro 14 M4 Max", ["1TB"], "14-inch M4 Max 1TB row"),
    ("2022 Macbook Air 13 inch M2", "MLY13 MLXW3 MLXY3 MLY33"): ("MacBook Air 13 M2", ["256GB"], "Part-code row inferred from standard 2022 M2 Air lineup"),
    ("2022 Macbook Pro 13 inch M2 NEW", "MC7U4 MC7V4 MC7W4 MC7X4"): ("MacBook Pro 13 M2", ["256GB"], "Part-code row inferred from standard 2022 M2 Pro lineup"),
    ("2020 MacBook Air 13", "MGN63LL/A MGND3LL/A MGN93LLA"): ("MacBook Air 13 M1", ["256GB"], "Part-code row inferred from standard 2020 M1 Air lineup"),
}


def parse_macbook_detail(header: str, detail: str) -> tuple[Optional[str], str, str, str, List[str], str]:
    header_norm = normalize_text(header)
    detail_norm = normalize_text(detail)
    if "macbook neo" in header_norm:
        return None, "", "", "", [], "Excluded MacBook Neo because it does not map cleanly to standard Apple MacBook naming."

    override = MACBOOK_OVERRIDES.get((normalize_space(header), normalize_space(detail)))
    if override:
        year_match = re.search(r"\b(20\d{2})\b", header_norm)
        year = year_match.group(1) if year_match else ""
        line = "macbook"
        if "macbook air" in normalize_text(override[0]):
            line = "macbook air"
        elif "macbook pro" in normalize_text(override[0]):
            line = "macbook pro"
        chip_match = re.search(r"\bm\s*([1-9])(?:\s+(pro|max))?\b", normalize_text(override[0]))
        chip_family = ""
        if chip_match:
            chip_family = f"m {chip_match.group(1)}"
            if chip_match.group(2):
                chip_family = f"{chip_family} {chip_match.group(2)}"
        return (
            override[0],
            year,
            line,
            chip_family,
            [normalize_storage_option(item) for item in override[1]],
            override[2],
        )

    line = "macbook"
    if "macbook air" in header_norm:
        line = "macbook air"
    elif "macbook pro" in header_norm:
        line = "macbook pro"
    else:
        return None, "", "", "", [], "Excluded non-MacBook row from MacBook sheet."

    year_match = re.search(r"\b(20\d{2})\b", header_norm)
    year = year_match.group(1) if year_match else ""

    size_match = re.search(r"(\d+)\s*inch", header_norm)
    size = size_match.group(1) if size_match else ""

    chip_match = re.search(r"\bm\s*([1-9])(?:\s+(pro|max))?\b", detail_norm or header_norm)
    chip = ""
    if chip_match:
        chip = f"m {chip_match.group(1)}"
        if chip_match.group(2):
            chip = f"{chip} {chip_match.group(2)}"
    ram_match = re.search(r"(\d+)\s*gb", detail_norm)
    ram = f"{ram_match.group(1)}gb" if ram_match else ""
    core_match = re.search(r"(\d+\s*c\s*-\s*\d+\s*c)", detail_norm)
    cores = core_match.group(1).replace(" ", "") if core_match else ""
    storage_match = re.search(r"(\d+\s*(?:gb|tb))", detail_norm)
    storage = normalize_storage_option(storage_match.group(1)) if storage_match else ""

    model_parts = [line]
    if size:
        model_parts.append(size)
    if chip:
        model_parts.append(chip)
    if ram:
        model_parts.append(ram)
    if cores:
        model_parts.append(cores)
    model = " ".join(part for part in model_parts if part)
    note = detail if detail else header
    return model, year, line, chip, [storage] if storage else [], note


def parse_macbook_rows(workbook) -> List[dict]:
    ws = workbook["MacBook Master"]
    rows = []
    current_header = ""
    current_detail = ""
    for row_idx in range(1, ws.max_row + 1):
        cell = ws[f"A{row_idx}"].value
        if not isinstance(cell, str):
            continue
        text = normalize_space(cell)
        sealed = as_number(ws[f"B{row_idx}"].value)
        opened = as_number(ws[f"C{row_idx}"].value)
        activated = as_number(ws[f"D{row_idx}"].value)

        if isinstance(ws[f"C{row_idx}"].value, str) and normalize_text(ws[f"C{row_idx}"].value) == "open":
            current_header = text
            current_detail = ""
            continue

        if any(value is not None and value > 0 for value in [sealed, opened, activated]):
            model, model_year, device_line, chip_family, storage_options, note = parse_macbook_detail(current_header, current_detail or text)
            if not model:
                continue
            condition_prices = {
                "sealed": sealed,
                "open": opened,
                "activated": activated,
            }
            condition_prices = {k: v for k, v in condition_prices.items() if v is not None and v > 0}
            default_condition, base_price = choose_default_condition(condition_prices, ["open", "activated", "sealed"])
            if base_price is None:
                continue
            rows.append(
                row_record(
                    category="macbook",
                    brand="apple",
                    model=model,
                    model_year=model_year,
                    device_line=device_line,
                    chip_family=chip_family,
                    storage_options=storage_options,
                    carrier_status="not_applicable",
                    default_condition=default_condition,
                    base_price_usd=base_price,
                    condition_prices_usd=condition_prices,
                    deduction_rules_usd={},
                    source_sheet="MacBook Master",
                    notes=note,
                )
            )
            continue

        lower = normalize_text(text)
        if "macbook" in lower:
            current_header = text
            current_detail = ""
        elif any(token in lower for token in ["gb", "tb", "chip", "/"]):
            current_detail = text

    return rows


def parse_watch_and_major_categories(workbook) -> List[dict]:
    rows = []
    rows.extend(parse_iphone_rows(workbook))
    rows.extend(parse_ipad_rows(workbook))
    rows.extend(parse_samsung_rows(workbook))
    rows.extend(parse_pixel_rows(workbook))
    rows.extend(parse_macbook_rows(workbook))
    rows.extend(parse_watch_rows(workbook))
    return merge_rows(rows)


def write_csv(rows: List[dict], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate prices.csv from the Atlas workbook")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Path to the Atlas workbook")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Path to write prices.csv")
    return parser.parse_args()


def main() -> None:
    from openpyxl import load_workbook

    args = parse_args()
    workbook_path = Path(args.workbook).expanduser()
    output_path = Path(args.output).expanduser()
    workbook = load_workbook(workbook_path, data_only=True)
    rows = parse_watch_and_major_categories(workbook)
    write_csv(rows, output_path)
    print(f"Wrote {len(rows)} rows to {output_path}")


if __name__ == "__main__":
    main()
